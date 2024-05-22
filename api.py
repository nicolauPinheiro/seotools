from flask import Flask, request, jsonify
import csv
import uuid
import numpy as np
from sklearn.cluster import AgglomerativeClustering
from sklearn.metrics.pairwise import cosine_similarity
from sentence_transformers import SentenceTransformer
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from io import StringIO
import os
from flask_cors import CORS  # Importar CORS
app = Flask(__name__)
CORS(app)  # Habilitar CORS
model = SentenceTransformer("sentence-transformers/paraphrase-xlm-r-multilingual-v1")

def get_embedding(keyword):
    keyword_embedding = model.encode([keyword])[0]
    return keyword_embedding

@app.route('/upload', methods=['POST'])
def upload_file():
    if 'file' not in request.files:
        return "No file part", 400
    
    file = request.files['file']
    threshold = float(request.form.get('threshold', 0.8))
    
    if file.filename == '':
        return "No selected file", 400
    
    if file and file.filename.endswith('.csv'):
        file_stream = StringIO(file.stream.read().decode("utf-8"))
        reader = csv.reader(file_stream)
        keywords = [row[0] for row in reader]
        
        # Generate embeddings
        keyword_embeddings = [get_embedding(keyword) for keyword in keywords]

        # Compute similarity matrix
        similarity_matrix = cosine_similarity(keyword_embeddings)

        # Perform clustering
        clustering = AgglomerativeClustering(n_clusters=None, metric="precomputed", linkage="average", distance_threshold=1-threshold)
        clusters = clustering.fit_predict(1 - similarity_matrix)
        randomico = str(uuid.uuid4())
        # Prepare Excel output
        output_filename = "keywords_output_"+randomico+".xlsx"
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Keywords"
        header = ["Removed Keyword", "Retained Keyword", "Similarity"]
        sheet.append(header)

        # Define color formatting
        yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        
        unique_keywords = []

        # Process clusters and write results to Excel
        for cluster_id in set(clusters):
            cluster_indices = np.where(clusters == cluster_id)[0]
            cluster_embeddings = [keyword_embeddings[i] for i in cluster_indices]
            centroid = np.mean(cluster_embeddings, axis=0)
            closest_index = min(cluster_indices, key=lambda i: np.linalg.norm(keyword_embeddings[i] - centroid))
            retained_keyword = keywords[closest_index]

            unique_keywords.append(retained_keyword)

            # Write removed keywords with yellow background
            for i in cluster_indices:
                if i != closest_index:
                    removed_keyword = keywords[i]
                    similarity = similarity_matrix[i, closest_index]
                    row = [removed_keyword, retained_keyword, similarity]
                    sheet.append(row)
                    for cell, fill in zip(sheet[sheet.max_row], [yellow_fill, yellow_fill, yellow_fill]):
                        cell.fill = fill

        # Save the Excel file
        workbook.save(output_filename)

        # Save unique keywords to a new CSV file
        unique_keywords_filename = "unique_keywords_"+randomico+".csv"
        with open(unique_keywords_filename, "w", newline='', encoding="utf-8") as f:
            writer = csv.writer(f)
            for keyword in unique_keywords:
                writer.writerow([keyword])

        response = {
            "message": "Processing completed successfully.",
            "output_excel_file": output_filename,
            "unique_keywords_file": unique_keywords_filename,
            "total_unique_keywords": len(unique_keywords),
            "total_removed_keywords": len(keywords) - len(unique_keywords)
        }
        return jsonify(response), 200

    return "Invalid file type", 400

if __name__ == "__main__":
    app.run(debug=True)
